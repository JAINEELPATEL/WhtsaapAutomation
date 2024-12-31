import time
import random
import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC



USER_EXCEL_FILE = "Attachments\\UserSheet.xlsx"
REPORT_EXCEL_FILE = "Attachments\\ReportSheet.xlsx"
MESSAGE_FILE = "Attachments\\MessageEnglish.txt"



CUSTOMER_DATA = {}
REPORT = []
MESSAGE = """"""



def excel_sheet_formatting():
    global CUSTOMER_DATA
    
    def format_date(cell_value):
        if isinstance(cell_value, datetime.datetime):
            return cell_value.strftime("%d-%m-%Y")
        return cell_value

    wb = openpyxl.load_workbook(USER_EXCEL_FILE, data_only=True)
    ws = wb.active
    
    header_row = None
    header_row_num = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and ("name" in str(cell).lower() or "number" in str(cell).lower()):
                header_row = row
                header_row_num = row_idx
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("No header row found containing 'name' or 'number' substring.")
            
    for header in header_row:
        if header is not None:
            CUSTOMER_DATA[header] = []

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        for idx, cell in enumerate(row):
            header = header_row[idx]
            if header is not None:
                formatted_cell = format_date(cell)
                CUSTOMER_DATA[header].append(formatted_cell)
    
    headers = list(CUSTOMER_DATA.keys())
    
    for key in headers:
        if key is not None:
            new_key = key.replace(".", "").replace(" ", "_")
            CUSTOMER_DATA[new_key] = CUSTOMER_DATA.pop(key)
        


def read_message():
    global MESSAGE

    with open(MESSAGE_FILE, "r") as file:
            MESSAGE = file.read()
    MESSAGE = MESSAGE.replace("{ ", "{").replace(" }", "}")
    
        

def chat_bot(sender_number):
    global CUSTOMER_DATA, MESSAGE, REPORT
    
    CHROME_DRIVER = webdriver.Chrome()
    CHROME_DRIVER.maximize_window()
    CHROME_DRIVER.get("https://web.whatsapp.com/")

    # Logic to login to sender account

    link_with_number = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//span[@role='button']"))
    )
    link_with_number.click()
    
    select_country = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//span[@class='xurb0ha']"))
    )
    select_country.click()
            
    enter_country_text_area = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//div[@role='textbox']//p"))
    )
    enter_country_text_area.send_keys("91")
    
    country_list = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='x1c4vz4f xs83m0k xdl72j9 x1g77sc7 x78zum5 xozqiw3 x1oa3qoh x12fk4p8 x1y1aw1k x1sxyh0 xwib8y2 xurb0ha xeuugli x2lwn1j x1nhvcw1 xdt5ytf x1qjc9v5']"))
    )
    country_list.send_keys(Keys.ARROW_DOWN, Keys.ENTER)

    enter_number_text_area = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//input[@value='+91 ']"))
    )
    enter_number_text_area.send_keys(sender_number)

    click_next = WebDriverWait(CHROME_DRIVER, 120).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Next')]"))
    )
    click_next.click()

    time.sleep(60)
    
    # Logic to find receiver and send the message
    
    num_records = len(next(iter(CUSTOMER_DATA.values())))
    
    for i, receiver_name, receiver_number in zip(range(num_records), CUSTOMER_DATA["NAME"], CUSTOMER_DATA["MOBILE_NO"]):
        values = {}
        for key in CUSTOMER_DATA:
            values[key] = CUSTOMER_DATA[key][i]
        
        message = MESSAGE.format(**values)

        search_user = WebDriverWait(CHROME_DRIVER, 120).until(
            EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true']"))
        )
        search_user.send_keys(receiver_number, Keys.ENTER)

        try: 
            text_message_area = WebDriverWait(CHROME_DRIVER, 2).until(
                EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Type a message']//p[@class='selectable-text copyable-text x15bjb6t x1n2onr6']"))
            )
            text_message_area.click()
            
            str = ""
            for char in message:
                if char == '\n':
                    text_message_area.send_keys(str, Keys.SHIFT, Keys.ENTER)
                    str = ""
                else:
                    str = str + char
            text_message_area.send_keys(Keys.ENTER)
        except TimeoutException:
            REPORT.append({"Name":receiver_name, "Phone No.":receiver_number})
                
        time.sleep(2)
        CHROME_DRIVER.refresh()
        time.sleep(random.randint(1, 5))
                    
    time.sleep(10)
    
    CHROME_DRIVER.quit()



def loading_unsend_data():
    wb = openpyxl.load_workbook(REPORT_EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            
    ws["A1"] = "Name"
    ws["B1"] = "Phone no."

    for index, entry in enumerate(REPORT, start=2):
        ws[f"A{index}"] = entry["Name"]
        ws[f"B{index}"] = entry["Phone No."]

    wb.save(REPORT_EXCEL_FILE)



def main():
    excel_sheet_formatting()
    read_message()
    chat_bot("9726337020")
    loading_unsend_data()



if __name__ == "__main__":
    main()
    