import os
import time
import random
import openpyxl
import datetime
import urllib.parse
import werkzeug.utils
from flask import Flask
from flask import request
from flask import jsonify
from flask_cors import CORS
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC




app = Flask(__name__)
CORS(app)
CORS(app, resources={r"/*": {"origins": "http://localhost:5174"}})



UPLOAD_FOLDER = "UPLOAD"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)



USER_EXCEL_FILE = None
REPORT_EXCEL_FILE = "C:\\Users\\BRIGHT\\Desktop\\__niikett__\\Automate WhatsApp\\Attachments\\ReportSheet.xlsx"
MESSAGE_FILE = ""



CUSTOMER_DATA = {}
REPORT = []
MESSAGE = ""



@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    global USER_EXCEL_FILE
    
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file and file.filename.endswith('.xlsx'):
        filename = werkzeug.utils.secure_filename(file.filename)
        USER_EXCEL_FILE = os.path.join(UPLOAD_FOLDER, filename)
        file.save(USER_EXCEL_FILE)

        USER_EXCEL_FILE = USER_EXCEL_FILE

        wb = openpyxl.load_workbook(USER_EXCEL_FILE)
        ws = wb.active

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        response_data = {
            "success": True,
            "message": "Excel file uploaded and processed successfully.",
            "data": data
        }

        return jsonify(response_data), 200
    else:
        return jsonify({"error": "Invalid file format. Please upload a .xlsx file."}), 400
    
    
      
@app.route('/excel-sheet-formatting', methods=['GET'])
def excel_sheet_formatting():
    # condition name should be save as NAME and number should be save as NUMBER
    global CUSTOMER_DATA
    
    def format_date(cell_value):
        if isinstance(cell_value, datetime.datetime):
            return cell_value.strftime("%d-%m-%Y")
        return cell_value

    wb = openpyxl.load_workbook(USER_EXCEL_FILE, data_only=True)
    ws = wb.active
    
    header_row = None
    header_row_num = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and ("name" or "number") in str(cell).lower():
                header_row = row
                header_row_num = row_idx
                break
        if header_row:
            break

    if header_row is None:
        return jsonify({"error": "No NAME and NUMBER found in excel sheet"}), 400
            
    for header in header_row:
        CUSTOMER_DATA[header] = []

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        for idx, cell in enumerate(row):
            header = header_row[idx]
            formatted_cell = format_date(cell)
            CUSTOMER_DATA[header].append(formatted_cell)
    
    headers = list(CUSTOMER_DATA.keys())
    
    for key in headers:
        if key is not None:
            new_key = key.replace(".", "").replace(" ", "_")
            CUSTOMER_DATA[new_key] = CUSTOMER_DATA.pop(key)
        
    headers = list(CUSTOMER_DATA.keys())

    return jsonify({"customer_data": CUSTOMER_DATA, "headers": headers})



@app.route('/write-message', methods=['POST'])
def write_message():
    global MESSAGE, MESSAGE_FILE
    
    data = request.json
    message = data.get('message', '')
    
    if not message:
        return jsonify({"error": "No message provided"}), 400
    
    directory = "messages"
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    MESSAGE_FILE = os.path.join(directory, "message.txt")
    
    with open(MESSAGE_FILE, "w") as file:
        file.write(message)
    
    with open(MESSAGE_FILE, "r") as file:
        MESSAGE = file.read()
    
    return jsonify({"message": "Message saved successfully", "file_path": MESSAGE_FILE})



def chat_bot():
    global CUSTOMER_DATA, MESSAGE, REPORT
        
    CHROME_DRIVER = webdriver.Chrome()
    CHROME_DRIVER.maximize_window()
    CHROME_DRIVER.get("https://web.whatsapp.com/")

    # Logic to login to sender account
    
    window = WebDriverWait(CHROME_DRIVER, 300).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='_al_c']"))
    )

    # Logic to find receiver and send the message
    
    num_records = len(next(iter(CUSTOMER_DATA.values())))
    
    for i, name, number in zip(range(num_records), CUSTOMER_DATA["NAME"], CUSTOMER_DATA["NUMBER"]):
        values = {}
        for key in CUSTOMER_DATA:
            values[key] = CUSTOMER_DATA[key][i]
        
        message = MESSAGE.format(**values)

        CHROME_DRIVER.minimize_window()

        link = f"https://web.whatsapp.com/send/?phone=91{number}&text={urllib.parse.quote(message)}"
        CHROME_DRIVER.get(link)
        
        CHROME_DRIVER.maximize_window()
            
        try:
            invalid_url = WebDriverWait(CHROME_DRIVER, 20).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'OK')]"))
            )
            
            # REPORT.append({"NAME":name, "NUMBER":number})
            with open("file.txt", "a") as file:
                file.write(f"{name} {number}\n")
            
            # with open(f"Message/{name}.txt", "w") as file:
            #     file.write(message)
        except Exception:
            action = ActionChains(CHROME_DRIVER)
            action.send_keys(Keys.ENTER)
            action.perform()
                
        time.sleep(2)
        CHROME_DRIVER.refresh()
        time.sleep(random.randint(1, 5))
                    
    time.sleep(10)
    
    CHROME_DRIVER.quit()
    return {"success": "Messages sent successfully"}



@app.route('/chat-bot', methods=['POST'])
def chat_bot_endpoint():
    data = request.get_json()
    sender_number = data.get('sender_number')
    
    if not sender_number:
        return jsonify({"error": "Sender number is required"}), 400

    try:
        result = chat_bot(sender_number)
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
        

@app.route('/loading-unsend-data', methods=['GET'])
def loading_unsend_data():
    global REPORT

    wb = openpyxl.load_workbook(REPORT_EXCEL_FILE)
    ws = wb.active
           
    ws["A1"] = "Name"
    ws["B1"] = "Phone no."

    for index, entry in enumerate(REPORT, start=2):
        ws[f"A{index}"] = entry["Name"]
        ws[f"B{index}"] = entry["Phone No."]

    wb.save(REPORT_EXCEL_FILE)
    return jsonify({"status": "success"})



if __name__ == "__main__":
    app.run(debug=True)
    