import time
import random
import openpyxl
import datetime
import urllib.parse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC

 

USER_EXCEL_FILE = "UserSheet.xlsx"
REPORT_EXCEL_FILE = "ReportSheet.xlsx"
MESSAGE_FILE = "Prompt\\Prompt3.txt"



CUSTOMER_DATA = {}
REPORT = []
MESSAGE = """"""



def excel_sheet_formatting():
    global CUSTOMER_DATA
    
    def format_date(cell_value):
        if isinstance(cell_value, datetime.datetime):
            return cell_value.strftime("%d-%m-%Y")
        return cell_value

    wb = openpyxl.load_workbook(USER_EXCEL_FILE, data_only=True)
    ws = wb.active
    
    header_row = None
    header_row_num = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and ("name" in str(cell).lower() or "number" in str(cell).lower()):
                header_row = row
                header_row_num = row_idx
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("No header row found containing 'name' or 'number' substring.")
            
    for header in header_row:
        if header is not None:
            CUSTOMER_DATA[header] = []

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        for idx, cell in enumerate(row):
            header = header_row[idx]
            if header is not None:
                formatted_cell = format_date(cell)
                CUSTOMER_DATA[header].append(formatted_cell)
    
    headers = list(CUSTOMER_DATA.keys())
    
    for key in headers:
        if key is not None:
            new_key = key.replace(".", "").replace(" ", "_")
            CUSTOMER_DATA[new_key] = CUSTOMER_DATA.pop(key)
        


def read_message():
    global MESSAGE

    with open(MESSAGE_FILE, "r", encoding = "utf-8") as file:
            MESSAGE = file.read()
    MESSAGE = MESSAGE.replace("{ ", "{").replace(" }", "}")
    
        

def chat_bot():
    global CUSTOMER_DATA, MESSAGE, REPORT
        
    CHROME_DRIVER = webdriver.Chrome()
    CHROME_DRIVER.maximize_window()
    CHROME_DRIVER.get("https://web.whatsapp.com/")

    # Logic to login to sender account
    
    window = WebDriverWait(CHROME_DRIVER, 300).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='_al_c']"))
    )

    # Logic to find receiver and send the message
    
    num_records = len(next(iter(CUSTOMER_DATA.values())))
    
    for i, name, number in zip(range(num_records), CUSTOMER_DATA["NAME"], CUSTOMER_DATA["NUMBER"]):
        values = {}
        for key in CUSTOMER_DATA:
            values[key] = CUSTOMER_DATA[key][i]
        
        message = MESSAGE.format(**values)

        # CHROME_DRIVER.minimize_window()

        link = f"https://web.whatsapp.com/send/?phone=91{number}&text={urllib.parse.quote(message)}"
        CHROME_DRIVER.get(link)
        
        # CHROME_DRIVER.maximize_window()
            
        try:
            invalid_url = WebDriverWait(CHROME_DRIVER, 20).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'OK')]"))
            )
            
            # REPORT.append({"NAME":name, "NUMBER":number})
            with open("file.txt", "a") as file:
                file.write(f"{name} {number}\n")
            
            # with open(f"Message/{name}.txt", "w") as file:
            #     file.write(message)
        except Exception:
            action = ActionChains(CHROME_DRIVER)
            action.send_keys(Keys.ENTER)
            action.perform()
                
        time.sleep(2)
        CHROME_DRIVER.refresh()
        time.sleep(random.randint(1, 5))
                    
    time.sleep(10)
    
    CHROME_DRIVER.quit()



def loading_unsend_data():
    global REPORT
    
    wb = openpyxl.load_workbook(REPORT_EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            
    ws["A1"] = "NAME"
    ws["B1"] = "NUMBER"

    for index, entry in enumerate(REPORT, start=2):
        ws[f"A{index}"] = entry["NAME"]
        ws[f"B{index}"] = entry["NUMBER"]

    wb.save(REPORT_EXCEL_FILE)



def main():
    excel_sheet_formatting()
    read_message()
    chat_bot()
    loading_unsend_data()



if __name__ == "__main__":
    main()
    