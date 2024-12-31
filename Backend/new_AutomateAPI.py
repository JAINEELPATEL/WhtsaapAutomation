import os
import time
import random
import openpyxl
import datetime
import urllib.parse
import uvicorn
from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_FOLDER = "UPLOAD"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

USER_EXCEL_FILE = None
REPORT_EXCEL_FILE = "C:\\Users\\BRIGHT\\Desktop\\__niikett__\\Automate WhatsApp\\Attachments\\ReportSheet.xlsx"
MESSAGE_FILE = ""
CUSTOMER_DATA = {}
REPORT = []
MESSAGE = ""

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    global USER_EXCEL_FILE, CUSTOMER_DATA

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload a .xlsx file.")

    USER_EXCEL_FILE = os.path.join(UPLOAD_FOLDER, file.filename)
    with open(USER_EXCEL_FILE, "wb") as f:
        f.write(await file.read())

    wb = openpyxl.load_workbook(USER_EXCEL_FILE)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    headers = excel_sheet_formatting()
    response_data = {
        "success": True,
        "message": "Excel file uploaded and processed successfully.",
        "data": data
    }
    return JSONResponse({"customer_data": CUSTOMER_DATA, "headers": headers, "response": response_data})

def excel_sheet_formatting():
    global CUSTOMER_DATA

    def format_date(cell_value):
        if isinstance(cell_value, datetime.datetime):
            return cell_value.strftime("%d-%m-%Y")
        return cell_value

    wb = openpyxl.load_workbook(USER_EXCEL_FILE, data_only=True)
    ws = wb.active

    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and ("name" in str(cell).lower() or "number" in str(cell).lower()):
                header_row = row
                break
        if header_row:
            break

    if header_row is None:
        raise HTTPException(status_code=400, detail="No NAME and NUMBER found in excel sheet")

    for header in header_row:
        CUSTOMER_DATA[header] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx, cell in enumerate(row):
            header = header_row[idx]
            formatted_cell = format_date(cell)
            CUSTOMER_DATA[header].append(formatted_cell)

    headers = list(CUSTOMER_DATA.keys())
    for key in headers:
        if key is not None:
            new_key = key.replace(".", "").replace(" ", "_")
            CUSTOMER_DATA[new_key] = CUSTOMER_DATA.pop(key)

    return list(CUSTOMER_DATA.keys())

@app.post("/write-message")
async def write_message(request: Request):
    global MESSAGE, MESSAGE_FILE

    data = await request.json()
    message = data.get("message")
    if not message:
        raise HTTPException(status_code=400, detail="No message provided")

    directory = "messages"
    if not os.path.exists(directory):
        os.makedirs(directory)

    MESSAGE_FILE = os.path.join(directory, "message.txt")
    with open(MESSAGE_FILE, "w") as file:
        file.write(message)

    MESSAGE = message
    return {"message": "Message saved successfully", "file_path": MESSAGE_FILE}

def chat_bot():
    global CUSTOMER_DATA, MESSAGE, REPORT

    CHROME_DRIVER = webdriver.Chrome()
    CHROME_DRIVER.maximize_window()
    CHROME_DRIVER.get("https://web.whatsapp.com/")

    WebDriverWait(CHROME_DRIVER, 300).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='_al_c']"))
    )

    num_records = len(next(iter(CUSTOMER_DATA.values())))
    for i, name, number in zip(range(num_records), CUSTOMER_DATA["NAME"], CUSTOMER_DATA["NUMBER"]):
        values = {key: CUSTOMER_DATA[key][i] for key in CUSTOMER_DATA}
        message = MESSAGE.format(**values)

        link = f"https://web.whatsapp.com/send/?phone=91{number}&text={urllib.parse.quote(message)}"
        CHROME_DRIVER.get(link)
        try:
            WebDriverWait(CHROME_DRIVER, 20).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'OK')]"))
            )
        except Exception:
            action = ActionChains(CHROME_DRIVER)
            action.send_keys(Keys.ENTER)
            action.perform()
        
        time.sleep(random.randint(1, 5))

    CHROME_DRIVER.quit()
    return {"success": "Messages sent successfully"}

@app.post("/chat-bot")
async def chat_bot_endpoint(request: Request):
    data = await request.json()
    sender_number = data.get("sender_number")

    if not sender_number:
        raise HTTPException(status_code=400, detail="Sender number is required")

    try:
        result = chat_bot()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/loading-unsend-data")
async def loading_unsend_data():
    global REPORT

    wb = openpyxl.load_workbook(REPORT_EXCEL_FILE)
    ws = wb.active

    ws["A1"] = "Name"
    ws["B1"] = "Phone no."

    for index, entry in enumerate(REPORT, start=2):
        ws[f"A{index}"] = entry["Name"]
        ws[f"B{index}"] = entry["Phone No."]

    wb.save(REPORT_EXCEL_FILE)
    return {"status": "success"}

if __name__ == "__main__":
    uvicorn.run("new_AutomateAPI:app", reload=True)
    